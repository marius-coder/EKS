# -*- coding: cp1252 -*-

import openpyxl
#open Workbook
wb_obj = openpyxl.load_workbook("./Ergebnis/Ergebnis_Dist.xlsx") 
#Read the Input sheet:
sheet = wb_obj["Eingabe"]
inputData = {
	"AutoDaten" : {
		"Anzahl Autos" : 0,
		"maxLadung" : 0,
		"Verbrauch" : 0,
		"Lade/Entladeleistung" : 0,
		"Effizienz" : 0,
		},	
	"PersonenDaten" : {
		"personenKilometer" : 0,
		"anzPersonen" : 0,
		"percentMobilität" : 0,
		"gfa" : 0,
		"percentAbweichung" : 0,
		"Safety" : 0,
		"Verbreitung externe Ladestationen" : 0
		},
	"ExterneDaten" : {
		"Info Lehrpersonal" : {
			"Ladung" : 0,
			"Anzahl" : 0,
			"Prozent Mitmachende" : 0,
			},
		"Info Sonstiges Personal" : {
			"Ladung" : 0,
			"Anzahl" : 0,
			"Prozent Mitmachende" : 0,
			}
		},
	"LadeDaten" : {	
		"distMinLadung" : {}
		}
	}

def SearchdistMinLadung():
	for row in sheet.iter_rows():
		if row[1].value == "distMinLadung":
			counter = 2
			ret = {}
			while(True):
				anteil = row[1].offset(counter,0).value
				if row[1].offset(counter,0).value == None:
					return ret
				ladung = row[1].offset(counter,1).value
				ret[str(anteil)] = ladung
				counter += 1


def SearchValue(name, special= ""):

	if name == "distMinLadung":
		ret = SearchdistMinLadung()
		return ret

	for row in sheet.iter_rows():
		if row[1].value == name:			
			if special == "gewerbe":
				return row[3].value
			else:
				return row[2].value


for gruppe in inputData.keys():
	for key in inputData[gruppe].keys():
		if type(inputData[gruppe][key]) == dict and inputData[gruppe][key]:
			for keyInner in inputData[gruppe][key].keys():
				if "Sonstiges" in key:
					inputData[gruppe][key][keyInner] = SearchValue(keyInner, special= "gewerbe")
				else:
					inputData[gruppe][key][keyInner] = SearchValue(keyInner)
		else:
			inputData[gruppe][key] = SearchValue(key)
